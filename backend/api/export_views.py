import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from django.http import HttpResponse

def generate_star_award_excel(nominations):
    """
    Logic to generate the Star Award Excel file.
    Accepts a queryset of Nominations and returns an HttpResponse with the Excel file.
    """
    
    # 1. Setup Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Star Award Export"

    # 2. Define EXACT Headers requested
    headers = [
        "Completion time",
        "Email",                                    
        "Name",                                     
        "Select an Award Type",                     
        "Enter the email address of the colleague you want to nominate", 
        "In which of the following categories should your nomination be included?",
        "Why are you nominating this colleague? Describe what your nominee did, the impact of their actions, who was affected, the expected duration of the impact, and any supporting feedback or evidence.",
        "Contract",
        "Location",
        "Country",
        "Practise",
        "Portfolio",
        "Line Manager",
        "Nomination Name",                          
        "Shortlist",                                
        "Successful mail back to nominators",       
        "Approaval-YES/NO"                          
    ]

    ws.append(headers)

    # 3. Styling Variables
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="008080", end_color="008080", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Apply styling to Header Row
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    # 4. Process Data
    for nom in nominations:      
        # A. Categories Parsing
        raw_metrics = nom.selected_metrics
        cat_list = []
        if isinstance(raw_metrics, str):
            try:
                data = json.loads(raw_metrics)
            except:
                data = []
        elif isinstance(raw_metrics, list):
            data = raw_metrics
        else:
            data = []

        for item in data:
            cat_list.append(item.get('category', ''))
        category_str = ", ".join(set(filter(None, cat_list)))

        # B. Helper to avoid empty cells (Returns "-" if empty)
        def get_val(obj, attr):
            val = getattr(obj, attr, '')
            return val if val else "-"

        # C. Determine Approval Status (Dynamic YES/NO)
        # YES = Coordinator Approved, Committee Approved, Awarded, or generic Approved
        # NO  = Submitted (pending), Coordinator Rejected, Committee Rejected
        approved_statuses = ['COORDINATOR_APPROVED', 'COMMITTEE_APPROVED', 'AWARDED', 'APPROVED']
        is_approved = nom.status in approved_statuses
        approval_text = "YES" if is_approved else "NO"

        # D. Build Row
        row = [
            nom.submitted_at.strftime("%Y-%m-%d %H:%M:%S"), # Completion time
            get_val(nom.nominator, 'email'),                # Nominator Email
            nom.nominator.username if nom.nominator else "System", # Nominator Name
            "Star Award",                                   # Award Type
            get_val(nom.nominee, 'email'),                  # Nominee Email
            category_str if category_str else "-",          # Category
            nom.reason if nom.reason else "-",              # Reason
            get_val(nom.nominee, 'contract_type'),          # Contract
            get_val(nom.nominee, 'location'),               # Location
            get_val(nom.nominee, 'country'),                # Country
            nom.nominee.employee_dept if nom.nominee.employee_dept else "-", # Practise
            nom.nominee.employee_role if nom.nominee.employee_role else "-", # Portfolio
            get_val(nom.nominee, 'line_manager_name'),      # Line Manager
            "Star Award",                                   # Nomination Name
            approval_text,                                  # Shortlist (Dynamic based on status)
            "YES",                                          # Successful mail back (Kept as YES)
            approval_text                                   # Approaval-YES/NO (Dynamic based on status)
        ]
        
        # Append Row
        ws.append(row)

        # This ensures the "-" or text is centered and wrapped
        for cell in ws[ws.max_row]:
            cell.alignment = center_align

    # 5. Formatting Widths (TIGHTER CAP)
    for column_cells in ws.columns:
        length = 0
        for cell in column_cells:
            if cell.row == 1:
                continue 
            if cell.value:
                length = max(length, len(str(cell.value)))
        final_width = min(max(length + 2, 15), 30)
        ws.column_dimensions[column_cells[0].column_letter].width = final_width

    # 6. Return Response
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="Star_Award_Export.xlsx"'
    wb.save(response)
    return response
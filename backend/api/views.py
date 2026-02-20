from rest_framework import generics, status, permissions
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework import permissions, status
from rest_framework_simplejwt.views import TokenObtainPairView
from rest_framework.pagination import PageNumberPagination
from django.contrib.auth import get_user_model
from django.db.models import Q, Count # Needed for search logic
from .models import Vote,Notification
from datetime import timedelta
from .export_views import generate_star_award_excel
from django.db.models import Count, F, Value
from openpyxl import Workbook
from django.db.models.functions import TruncDate, TruncMonth
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
from django.http import HttpResponse
from .models import Nomination, User  # Ensure User is imported
from .serializers import AdminVoteResultSerializer,NotificationSerializer 
from django.utils import timezone
from .ai_utils import get_nomination_sentiment
from .serializers import (
    UserRegistrationSerializer,
    UserProfileSerializer,
    UserNominationListSerializer,
    NominationSerializer,
    CustomLoginSerializer,
    AdminVoteResultSerializer,
    FinalistSerializer,
)
from .models import Nomination, NOMINATION_CRITERIA
from .utils import send_notification  
import openpyxl
from rest_framework.parsers import MultiPartParser, FormParser
User = get_user_model()
 
class RegisterView(generics.CreateAPIView):
    queryset = User.objects.all()
    permission_classes = (permissions.AllowAny,)
    serializer_class = UserRegistrationSerializer
 
class CustomLoginView(TokenObtainPairView):
    serializer_class = CustomLoginSerializer    
class UserProfileView(generics.RetrieveUpdateAPIView):
    permission_classes = [permissions.IsAuthenticated]
    serializer_class = UserProfileSerializer
 
    def get_object(self):
        return self.request.user      

# 1. Define the Pagination Class (Standard Practice)
class StandardResultsSetPagination(PageNumberPagination):
    page_size = 15  # Matches your frontend ITEMS_PER_PAGE
    page_size_query_param = 'page_size'
    max_page_size = 100


class NominationFilterOptionsView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        data = User.objects.exclude(role='ADMIN').values(
            'employee_dept', 
            'employee_role', 
            'location'
        ).distinct()
        
        return Response(list(data))
     
class NominationOptionsView(generics.ListAPIView):
    permission_classes = [permissions.IsAuthenticated]
    serializer_class = UserNominationListSerializer
    pagination_class = StandardResultsSetPagination 

    def get_queryset(self):
        user = self.request.user
        queryset = User.objects.exclude(id=user.id).exclude(role='ADMIN').order_by('username') # 🔥 Always order paginated lists!

        # 1. Unified Search
        search_query = self.request.query_params.get('search', None)
        if search_query:
            queryset = queryset.filter(
                Q(username__icontains=search_query) |
                Q(employee_id__icontains=search_query)
            )

        # 2. Filters
        dept_filter = self.request.query_params.get('dept', None)
        if dept_filter and dept_filter != "All":
            queryset = queryset.filter(employee_dept__iexact=dept_filter)

        role_filter = self.request.query_params.get('role', None)
        if role_filter and role_filter != "All":
            queryset = queryset.filter(employee_role__iexact=role_filter)

        loc_filter = self.request.query_params.get('location', None)
        if loc_filter and loc_filter != "All":
            queryset = queryset.filter(location__iexact=loc_filter)

        return queryset 
     
def check_timeline_validity(phase):
    return True, "Allowed"

class CreateNominationView(APIView):
    permission_classes = [permissions.IsAuthenticated]
 
    def post(self, request):
        # Time Check
        is_valid, msg = check_timeline_validity('NOMINATION')
        if not is_valid:
            return Response({"error": msg}, status=status.HTTP_403_FORBIDDEN)
 
        # If user has a nomination that is NOT Rejected (in any form), block them.
        existing_noms = Nomination.objects.filter(nominator=request.user).exclude(
            status__in=['COORDINATOR_REJECTED', 'REJECTED']
        )

        if existing_noms.exists():
            return Response(
                {"error": "You have already nominated someone. You can only nominate 1 person."},
                status=status.HTTP_400_BAD_REQUEST
            )
 
        serializer = NominationSerializer(data=request.data, context={'request': request})
       
        if serializer.is_valid():
            serializer.save(nominator=request.user)
            return Response({"message": "Nomination submitted successfully!"}, status=status.HTTP_201_CREATED)
       
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
class NominationStatusView(APIView):
    permission_classes = [permissions.IsAuthenticated]
 
    def get(self, request):
        user = request.user

        my_nomination = Nomination.objects.filter(nominator=user).exclude(
            status__in=['COORDINATOR_REJECTED', 'REJECTED']
        ).first()
       
        received_count = Nomination.objects.filter(nominee=user).count()
        nominee_data = None
        current_reason = None
        
        if my_nomination:
            nominee_data = UserNominationListSerializer(my_nomination.nominee).data
            current_reason = my_nomination.reason
 
        return Response({
            "has_nominated": my_nomination is not None, 
            "nominee": nominee_data,
            "nominee_name": my_nomination.nominee.username if my_nomination else None,
            "reason": current_reason,
            "nominee_id": my_nomination.nominee.id if my_nomination else None,
            "nomination_date": my_nomination.submitted_at if my_nomination else None,
            "nominations_received_count": received_count
        })
  
EDIT_WINDOW_DAYS = 2 
 
class NominationOptionsDataView(APIView):
    permission_classes = [permissions.IsAuthenticated]
 
    def get(self, request):
        # Sends the structure { "Category": ["Metric1", "Metric2"], ... }
        return Response(NOMINATION_CRITERIA) 
    
class ManageNominationView(APIView):
    permission_classes = [permissions.IsAuthenticated]
 
    def is_locked(self, nomination):
        # Locked if not in initial submitted state
        return nomination.status != 'NOMINATION_SUBMITTED'
 
    def get_my_nomination(self, user):
        # Get the latest non-coordinator-rejected nomination
        return Nomination.objects.filter(nominator=user).exclude(status='COORDINATOR_REJECTED').order_by('-submitted_at').first()
 
    #  CREATE NOMINATION
    def post(self, request):
        if Nomination.objects.filter(nominator=request.user).exclude(status='COORDINATOR_REJECTED').exists():
            return Response(
                {"error": "You have already nominated someone."},
                status=status.HTTP_400_BAD_REQUEST
            )
 
        serializer = NominationSerializer(
            data=request.data,
            context={'request': request}
        )
 
        if serializer.is_valid():
            nomination = serializer.save(nominator=request.user)
            
            send_notification(
                user=request.user,
                title="Nomination Confirmed",
                message=f"Hi {request.user.username}, thank you for nominating {nomination.nominee.username}. Your submission has been received.",
                notif_type="NOMINATION"
            )

            return Response({"message": "Nomination submitted successfully!"}, status=status.HTTP_201_CREATED)
 
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
 
    # UPDATE
    def put(self, request):
        nomination = self.get_my_nomination(request.user)
        if not nomination:
            return Response({"error": "No nomination found to edit."}, status=404)
 
        if self.is_locked(nomination):
            return Response(
                {"error": "This nomination has already been reviewed by an authority and cannot be edited."},
                status=403
            )
 
        serializer = NominationSerializer(
            nomination,
            data=request.data,
            context={'request': request},
            partial=True
        )
 
        if serializer.is_valid():
            serializer.save() 
            return Response({"message": "Nomination updated successfully!"})
 
        return Response(serializer.errors, status=400)
 
    # DELETE
    def delete(self, request):
        nomination = self.get_my_nomination(request.user)
        if not nomination:
            return Response({"error": "No nomination found."}, status=404)
 
        if self.is_locked(nomination):
            return Response(
                {"error": "Cannot withdraw. This nomination has already been reviewed by an authority."},
                status=403
            )
 
        nomination.delete()
        return Response({"message": "Nomination withdrawn successfully."})
    
class CoordinatorNominationView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        filter_type = request.query_params.get("filter", "pending")
        
        query = Nomination.objects.select_related("nominee", "nominator").order_by("-submitted_at")

        # 1. PENDING (Coordinator has not acted yet)
        if filter_type == "coordinator_pending" or filter_type == "pending":
            nominations = query.filter(status__in=["NOMINATION_SUBMITTED", "SUBMITTED", "Pending"])

        # 2. COMMITTEE PENDING (Coordinator Approved, waiting for Committee)
        elif filter_type == "committee_pending":
            nominations = query.filter(status__in=["COORDINATOR_APPROVED", "APPROVED"])

        # 3. HISTORY (Processed items)
        elif filter_type == "history":
            nominations = query.exclude(
                status__in=["NOMINATION_SUBMITTED", "SUBMITTED", "Pending"]
            )
            
        else:
            nominations = query.filter(status__in=["NOMINATION_SUBMITTED", "SUBMITTED"])

        data = []
        for n in nominations:
            #  LOGIC: Extract Category from selected_metrics JSON
            derived_category = "N/A"
            metrics_data = n.selected_metrics

            # Case A: It's already a Python List (Django JSONField default)
            if isinstance(metrics_data, list) and len(metrics_data) > 0:
                first_item = metrics_data[0]
                if isinstance(first_item, dict):
                    derived_category = first_item.get('category', 'N/A')

            # Case B: It's a String (sometimes happens with SQLite/CSV imports)
            elif isinstance(metrics_data, str):
                try:
                    parsed_data = json.loads(metrics_data)
                    if isinstance(parsed_data, list) and len(parsed_data) > 0:
                        derived_category = parsed_data[0].get('category', 'N/A')
                except:
                    derived_category = "N/A"

            data.append({
                "id": n.id,
                "nominee_name": f"{n.nominee.first_name} {n.nominee.last_name}".strip() or n.nominee.username,
                "nominee_role": getattr(n.nominee, 'employee_role', "Employee"),
                "nominee_dept": getattr(n.nominee, 'employee_dept', "General"),
                "nominator_name": f"{n.nominator.first_name} {n.nominator.last_name}".strip() or n.nominator.username,
                "reason": n.reason,
                "submitted_at": n.submitted_at,
                "status": n.status,
                
                #  SEND DERIVED CATEGORY
                "category": derived_category,
                "selected_metrics": n.selected_metrics
            })

        return Response(data)

    def post(self, request):
        nom_id = request.data.get("nomination_id")
        action = request.data.get("action") 
        user = request.user

        if user.role not in ["COORDINATOR", "ADMIN"]:
            return Response({"error": "Unauthorized."}, status=403)

        try:
            original_nom = Nomination.objects.get(id=nom_id)
            nominee = original_nom.nominee
            all_noms_for_person = Nomination.objects.filter(nominee=nominee)
        except Nomination.DoesNotExist:
            return Response({"error": "Nomination not found"}, status=404)

        #  REJECT LOGIC 
        if action == "REJECT":
            if original_nom.status in ["NOMINATION_SUBMITTED", "SUBMITTED", "Pending"]:
                new_status = "COORDINATOR_REJECTED"
                msg = "Rejected (Coordinator Phase)"
            elif original_nom.status in ["COORDINATOR_APPROVED", "APPROVED"]:
                new_status = "COMMITTEE_REJECTED"
                msg = "Rejected (Committee Phase)"
            else:
                return Response({"error": f"Cannot reject from current state: {original_nom.status}"}, status=400)

            for nom in all_noms_for_person:
                send_notification(
                    user=nom.nominator,
                    title="Nomination Update: Action Required",
                    message=(
                        f"Hi {nom.nominator.first_name or nom.nominator.username}, "
                        f"your nomination for {nominee.username} has been reviewed and was not selected to move forward at this time. "
                        f"You are encouraged to submit a new nomination with a more detailed reason, or you may nominate another deserving colleague."
                    ),
                    notif_type="INFO"
                )

            all_noms_for_person.update(status=new_status)
            return Response({"message": f"{msg} for {nominee.username}"})

        # APPROVE LOGIC 
        elif action == "APPROVE":
            if original_nom.status in ["NOMINATION_SUBMITTED", "SUBMITTED", "Pending"]:
                new_status = "COORDINATOR_APPROVED"
                msg = "Shortlisted by Coordinator"
            
            elif original_nom.status in ["COORDINATOR_APPROVED", "APPROVED"]:
                finalist_count = Nomination.objects.filter(status="COMMITTEE_APPROVED").values('nominee').distinct().count()
                if finalist_count >= 15:
                    return Response({"error": "Finalist limit (15) reached."}, status=400)
                
                new_status = "COMMITTEE_APPROVED"
                msg = "Approved by Committee (Finalist)"
            
            elif original_nom.status == "COMMITTEE_APPROVED":
                new_status = "AWARDED"
                msg = "Award Granted"
            else:
                return Response({"error": f"Cannot approve from current state: {original_nom.status}"}, status=400)

            send_notification(
                user=nominee,
                title="Congratulations! Your Nomination was Approved",
                message=(
                    f"Hi {nominee.first_name or nominee.username}, "
                    f"great news! A nomination submitted for you has been reviewed and approved "
                    f"Keep up the excellent work!"
                ),
                notif_type="INFO"
            )

            all_noms_for_person.update(status=new_status)
            return Response({"message": f"{msg} for {nominee.username}"})

        # UNDO LOGIC 
        elif action == "UNDO":
            reversion_map = {
                "AWARDED": "COMMITTEE_APPROVED",
                "COMMITTEE_APPROVED": "COORDINATOR_APPROVED",
                "COORDINATOR_APPROVED": "NOMINATION_SUBMITTED",
                "APPROVED": "NOMINATION_SUBMITTED",
                "COORDINATOR_REJECTED": "NOMINATION_SUBMITTED",
                "REJECTED": "NOMINATION_SUBMITTED",
                "COMMITTEE_REJECTED": "COORDINATOR_APPROVED"
            }
            
            new_status = reversion_map.get(original_nom.status)
            if not new_status:
                return Response({"error": "Cannot undo from this stage"}, status=400)
            
            all_noms_for_person.update(status=new_status)
            return Response({"message": f"Undo successful. Reverted to {new_status}"})

        return Response({"error": "Invalid Action"}, status=400)
    
class NotificationListView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        notifications = request.user.notifications.all()
        serializer = NotificationSerializer(notifications, many=True)
        return Response(serializer.data)

class NotificationMarkReadView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request, pk):
        try:
            notif = Notification.objects.get(id=pk, user=request.user)
            notif.is_read = True
            notif.save()
            return Response({"message": "Notification marked as read"})
        except Notification.DoesNotExist:
            return Response({"error": "Notification not found"}, status=404)
           
class VotingView(APIView):
    permission_classes = [permissions.IsAuthenticated]
 
    def get(self, request):
        has_voted = Vote.objects.filter(voter=request.user).exists()
        finalists = Nomination.objects.filter(status="COMMITTEE_APPROVED")
 
        unique = {}
        for r in finalists:
            if r.nominee.id not in unique:
                unique[r.nominee.id] = r
        finalists = list(unique.values())
 
        return Response({
            "has_voted": has_voted,
            "finalists": FinalistSerializer(finalists, many=True).data
        })
 
    def post(self, request):
        if request.user.role == "ADMIN": return Response({"error": "Admins cannot vote."}, status=403)
        if Vote.objects.filter(voter=request.user).exists(): return Response({"error": "You already voted."}, status=400)
        
        nom_id = request.data.get("nomination_id")
        try:
            nom = Nomination.objects.get(id=nom_id, status="COMMITTEE_APPROVED")
            Vote.objects.create(voter=request.user, nomination=nom)
            return Response({"message": "Vote submitted!"})
        except Nomination.DoesNotExist:
            return Response({"error": "Invalid finalist selected."}, status=404)
        
class AdminResultsView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        if request.user.role != 'ADMIN' and request.user.role != 'COORDINATOR':
            return Response({"error": "Unauthorized"}, status=403)

        results = Nomination.objects.filter(
            status__in=['COMMITTEE_APPROVED', 'AWARDED']
        ).select_related('nominee').annotate(
            vote_count=Count('votes')
        ).order_by('-vote_count')
        
        unique = {}
        for r in results:
            if r.nominee.id not in unique:
                unique[r.nominee.id] = r
    
        final_results = list(unique.values())

        data = []
        for nom in final_results:
            data.append({
                "id": nom.id,
                "nominee_name": f"{nom.nominee.first_name} {nom.nominee.last_name}".strip() or nom.nominee.username,
                "employee_id": nom.nominee.employee_id,
                "employee_role": nom.nominee.employee_role or "No Title",
                "employee_dept": nom.nominee.employee_dept or "General",
                "reason": nom.reason,
                "status": nom.status,
                "vote_count": nom.vote_count
            })
        
        return Response(data)

    def post(self, request):
        if request.user.role != 'ADMIN' and request.user.role != 'COORDINATOR':
            return Response({"error": "Unauthorized"}, status=403)

        winner_id = request.data.get('nomination_id')
        try:
            winner = Nomination.objects.get(id=winner_id)
            winner.status = 'AWARDED'
            winner.save()
            return Response({"message": "Winner declared!"})
        except Nomination.DoesNotExist:
            return Response({"error": "Nomination not found"}, status=404)
        
class WinnersView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        if request.user.role != 'ADMIN' and request.user.role != 'COORDINATOR':
             return Response({"error": "Unauthorized"}, status=403)  
        
        coordinator_winners = Nomination.objects.filter(
            status__in=['COORDINATOR_APPROVED', 'COMMITTEE_APPROVED', 'AWARDED']
        )    
        committee_winners = Nomination.objects.filter(
            status__in=['COMMITTEE_APPROVED', 'AWARDED']
        )  

        final_winners = Nomination.objects.filter(status='AWARDED')  

        def dedupe(queryset):
            unique = {}
            for n in queryset:
                if n.nominee.id not in unique:
                    unique[n.nominee.id] = n
            return list(unique.values())  

        coordinator_winners = dedupe(coordinator_winners)
        committee_winners = dedupe(committee_winners)  
        
        # ADDED: Deduplicate final winners in case of multiple nominations for same person
        final_winners = dedupe(final_winners)

        def serialize_nom(n):
            if not n: return None
            return {
                "username": f"{n.nominee.first_name} {n.nominee.last_name}".strip() or n.nominee.username,
                "employee_id": getattr(n.nominee, 'employee_id', n.nominee.id),
                "employee_role": n.nominee.employee_role,
                "employee_dept": n.nominee.employee_dept,
            }  

        return Response({
            # CHANGED: Return as an array mapped through serialize_nom
            "final_winners": [serialize_nom(n) for n in final_winners],
            "committee_winners": [serialize_nom(n) for n in committee_winners],
            "coordinator_winners": [serialize_nom(n) for n in coordinator_winners],
        })

class AdminAnalyticsView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        if request.user.role != 'ADMIN' and request.user.role != 'COORDINATOR':
            return Response({"error": "Unauthorized"}, status=403)

        total_nominations = Nomination.objects.count()

        # Coordinator Approved: Anyone who passed the first stage (including if failed later or won)
        coordinator_approved = Nomination.objects.filter(
            status__in=[
                'COORDINATOR_APPROVED',
                'COMMITTEE_APPROVED',
                'AWARDED',
                'COMMITTEE_REJECTED' # They were approved by coordinator, then rejected by committee
            ]
        ).count()

        # Total Rejections (Sum of both types)
        total_rejections = Nomination.objects.filter(
            status__in=['COORDINATOR_REJECTED', 'COMMITTEE_REJECTED']
        ).count()
        
        total_employees = User.objects.filter(role='EMPLOYEE').count()
        employees_who_nominated = Nomination.objects.values('nominator').distinct().count()
        employees_not_nominated = total_employees - employees_who_nominated

        committee_finalists = Nomination.objects.filter(
            status__in=['COMMITTEE_APPROVED', 'AWARDED']
        ).count()

        final_winner = (Nomination.objects.filter(status='AWARDED').values('nominee').distinct().count())

        # Department Stats
        dept_stats = (
            Nomination.objects
            .values('nominee__employee_dept')
            .annotate(count=Count('id'))
            .order_by('-count')
        )
        department_stats = [
            {"department": d['nominee__employee_dept'] or "Unknown", "count": d['count']}
            for d in dept_stats
        ]

        # Daily Trend
        daily_trend = (
            Nomination.objects
            .annotate(day=TruncDate("submitted_at"))
            .values("day")
            .annotate(count=Count("id"))
            .order_by("day")
        )
        daily_trend_data = [{"date": d["day"], "count": d["count"]} for d in daily_trend]

        # Monthly Trend
        monthly_trend = (
            Nomination.objects
            .annotate(month=TruncMonth("submitted_at"))
            .values("month")
            .annotate(count=Count("id"))
            .order_by("month")
        )
        trend_data = [{"month": m["month"].strftime("%Y-%m") if m["month"] else "Unknown", "count": m["count"]} for m in monthly_trend]

        return Response({
            "summary": {
                "total_nominations": total_nominations,
                "coordinator_approved": coordinator_approved,
                "committee_finalists": committee_finalists,
                "final_winner": final_winner,
                "total_rejections": total_rejections,
                "employees_not_nominated": employees_not_nominated,
            },
            "department_stats": department_stats,
            "daily_trend": daily_trend_data,
            "trend_data": trend_data
        })

# 9. REPORTS EXPORT - UPDATED
class AdminReportExportView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        if request.user.role != "ADMIN" and request.user.role != 'COORDINATOR':
            return Response({"error": "Unauthorized"}, status=403)

        wb = Workbook()
        ws = wb.active
        ws.title = "Summary"

        ws.append(["Metric", "Value"])
        ws.append(["Total Nominations", Nomination.objects.count()])
        ws.append([
            "Coordinator Approved",
            Nomination.objects.filter(
                status__in=["COORDINATOR_APPROVED", "COMMITTEE_APPROVED", "AWARDED", "COMMITTEE_REJECTED"]
            ).count()
        ])
        ws.append([
            "Committee Finalists",
            Nomination.objects.filter(status="COMMITTEE_APPROVED").count()
        ])
        ws.append([
            "Final Winners",
            Nomination.objects.filter(status="AWARDED")
            .values("nominee").distinct().count()
        ])
        ws.append([
            "Total Rejections",
            Nomination.objects.filter(status__in=["COORDINATOR_REJECTED", "COMMITTEE_REJECTED"]).count()
        ])
        
        total_employees = User.objects.filter(role='EMPLOYEE').count() 
        employees_who_nominated = Nomination.objects.values("nominator").distinct().count()
        employees_not_nominated = total_employees - employees_who_nominated

        ws.append(["Employees Not Nominated", employees_not_nominated])

        # Sheet 2: Dept Analytics (Same logic, code omitted for brevity but logic stands)
        ws2 = wb.create_sheet(title="Department Analytics")
        ws2.append(["Department", "Nomination Count"])
        department_analytics = Nomination.objects.values(department=F("nominee__employee_dept")).annotate(count=Count("id"))
        for d in department_analytics:
            ws2.append([d["department"], d["count"]])

        # Sheet 3: Logs
        ws3 = wb.create_sheet(title="Approval Logs")
        ws3.append(["Employee", "Department", "Stage", "Action By", "Date"])

        nominations = Nomination.objects.select_related("nominee", "nominator")

        for n in nominations:
            ws3.append([
                n.nominee.username,
                n.nominee.employee_dept,
                "Initial Nomination",
                n.nominator.username if n.nominator else "System",
                n.submitted_at.strftime("%Y-%m-%d"),
            ])

            # Coordinator Approval
            if n.status in ["COORDINATOR_APPROVED", "COMMITTEE_APPROVED", "AWARDED", "COMMITTEE_REJECTED"]:
                ws3.append([
                    n.nominee.username,
                    n.nominee.employee_dept,
                    "Coordinator Approval",
                    "Coordinator", 
                    n.submitted_at.strftime("%Y-%m-%d"),
                ])

            # Committee Approval
            if n.status in ["COMMITTEE_APPROVED", "AWARDED"]:
                ws3.append([
                    n.nominee.username,
                    n.nominee.employee_dept,
                    "Committee Approval",
                    "Committee",
                    n.submitted_at.strftime("%Y-%m-%d"),
                ])

            # Award
            if n.status == "AWARDED":
                ws3.append([
                    n.nominee.username,
                    n.nominee.employee_dept,
                    "Final Winner",
                    "Coordinator/Admin",
                    n.submitted_at.strftime("%Y-%m-%d"),
                ])

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="admin_report.xlsx"'
        wb.save(response)
        return response

# 10. AI ANALYSIS VIEW - UPDATED
class NominationAIAnalysisView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        # AI looks at SUBMITTED nominations
        nominations = Nomination.objects.filter(status="NOMINATION_SUBMITTED").select_related('nominee')

        if not nominations.exists():
            return Response([], status=200)

        grouped_data = {}

        for n in nominations:
            nominee_id = n.nominee.id
            raw_data = n.selected_metrics
            if isinstance(raw_data, str):
                try:
                    metrics_list = json.loads(raw_data)
                except:
                    metrics_list = []
            else:
                metrics_list = raw_data or []

            categories = {item.get('category', 'General') for item in metrics_list}
            metrics = {item.get('metric', 'Performance') for item in metrics_list}
            
            cat_str = ", ".join(categories)
            met_str = ", ".join(metrics)
            reason = n.reason or "No specific reason provided."

            if nominee_id not in grouped_data:
                grouped_data[nominee_id] = {
                    "first_nomination_id": n.id,
                    "name": n.nominee.username,
                    "email": getattr(n.nominee, 'email', 'N/A'),
                    "details": [],
                    "count": 0
                }
            
            detail_str = f"Focus: {cat_str} ({met_str}) -> {reason}"
            grouped_data[nominee_id]["details"].append(detail_str)
            grouped_data[nominee_id]["count"] += 1

        data_for_ai = []
        for nominee_id, data in grouped_data.items():
            combined_text = " | ".join(data["details"])
            full_prompt_text = f"Candidate: {data['name']}. Received {data['count']} nominations. Inputs: {combined_text}"
            data_for_ai.append({
                "id": data["first_nomination_id"], 
                "reason": full_prompt_text
            })

        ai_results = get_nomination_sentiment(data_for_ai)
        ai_lookup = {item['id']: item for item in ai_results}
        final_response = []

        for nominee_id, data in grouped_data.items():
            lookup_id = data["first_nomination_id"]
            ai_data = ai_lookup.get(lookup_id, {})

            final_response.append({
                "id": lookup_id,
                "name": data["name"],
                "email": data["email"],
                "votes": data["count"], 
                "summary": ai_data.get('summary', "Analysis Pending..."),
                "sentiment": ai_data.get('sentiment', "Neutral"),
                "status": "SUBMITTED"
            })

        return Response(final_response, status=status.HTTP_200_OK)
    
class UserManagementView(APIView):
    # Support both File Uploads (Multipart) and JSON (Manual Entry)
    parser_classes = (MultiPartParser, FormParser, JSONParser)
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request, *args, **kwargs):
        # CASE 1: BULK UPLOAD (File Present)
        if 'file' in request.FILES:
            file_obj = request.FILES['file']
            try:
                wb = openpyxl.load_workbook(file_obj)
                sheet = wb.active
                
                created_count = 0
                updated_count = 0
                
                # Helper to safely get cell values
                def get_val(row, idx):
                    try:
                        val = row[idx]
                        return str(val).strip() if val is not None else ""
                    except IndexError:
                        return ""

                # Skip header, iterate rows
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    email = get_val(row, 7) # Column H (Email)
                    if not email: continue

                    # Map Excel Columns
                    full_name = get_val(row, 6)
                    first_name, last_name = self.split_name(full_name)
                    
                    user_data = {
                        'contract_type': get_val(row, 0),
                        'location': get_val(row, 1),
                        'country': get_val(row, 2),
                        'practice': get_val(row, 3),
                        'portfolio': get_val(row, 4),
                        'line_manager_name': get_val(row, 5),
                        'employee_dept': get_val(row, 3), 
                    }
                    
                    if self.save_user(email, first_name, last_name, user_data):
                        created_count += 1
                    else:
                        updated_count += 1

                return Response({
                    "message": "Bulk upload complete",
                    "mode": "bulk",
                    "created": created_count,
                    "updated": updated_count
                }, status=status.HTTP_200_OK)

            except Exception as e:
                return Response({"error": f"Excel Error: {str(e)}"}, status=500)

        # CASE 2: SINGLE USER (No File, just Data)

        elif 'email' in request.data:
            data = request.data
            email = data.get('email')
            full_name = data.get('name', '')
            first_name, last_name = self.split_name(full_name)

            user_data = {
                'contract_type': data.get('contract_type', ''),
                'location': data.get('location', ''),
                'country': data.get('country', ''),
                'practice': data.get('practice', ''),
                'portfolio': data.get('portfolio', ''),
                'line_manager_name': data.get('line_manager_name', ''),
                'employee_dept': data.get('practice', ''),
                'employee_id': data.get('employee_id', ''),
            }

            is_created = self.save_user(email, first_name, last_name, user_data)
            
            msg = "User created successfully" if is_created else "User updated successfully"
            return Response({"message": msg, "mode": "single"}, status=200)

        # CASE 3: INVALID REQUEST
        else:
            return Response({"error": "Provide either a 'file' or 'email' and 'name'"}, status=400)

    def split_name(self, full_name):
        if not full_name: return "", ""
        parts = full_name.strip().split(' ')
        return parts[0], " ".join(parts[1:]) if len(parts) > 1 else ""

    def save_user(self, email, first_name, last_name, extra_data):
        # 1. Update Username Logic: "mansi.l@..." -> "mansi"
        # Take the part before '@', then split by '.' and take the first part
        email_prefix = email.split('@')[0]
        username = email_prefix.split('.')[0]

        # 2. Update Employee ID Logic: Strictly None if missing/empty
        emp_id = extra_data.get('employee_id')
        if not emp_id or str(emp_id).strip() == "":
            emp_id = None 

        # 3. Update/Create User
        user, created = User.objects.update_or_create(
            email=email,
            defaults={
                'username': username,
                'first_name': first_name,
                'last_name': last_name,
                'employee_id': emp_id,  
                
                'contract_type': extra_data.get('contract_type'),
                'location': extra_data.get('location'),
                'country': extra_data.get('country'),
                'line_manager_name': extra_data.get('line_manager_name'),
                
                
                'employee_dept': extra_data.get('practice'),   # Stores Practice
                'employee_role': extra_data.get('portfolio'),  # Stores Portfolio
            }
        )
        
        if created:
            user.set_password(username)
            user.save()
        
        return created

# 10. AI ANALYSIS VIEW - UPDATED (Fixes "No Data" in Co-pilot)
class NominationAIAnalysisView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        # FIX: Look for BOTH 'NOMINATION_SUBMITTED' 
        nominations = Nomination.objects.filter(
            status__in=["NOMINATION_SUBMITTED", "SUBMITTED"]
        ).select_related('nominee')

        if not nominations.exists():
            return Response([], status=200)

        grouped_data = {}

        for n in nominations:
            nominee_id = n.nominee.id
            
            # Safe JSON parsing
            raw_data = n.selected_metrics
            if isinstance(raw_data, str):
                try:
                    metrics_list = json.loads(raw_data)
                except:
                    metrics_list = []
            else:
                metrics_list = raw_data or []

            categories = {item.get('category', 'General') for item in metrics_list}
            metrics = {item.get('metric', 'Performance') for item in metrics_list}
            
            cat_str = ", ".join(categories)
            met_str = ", ".join(metrics)
            reason = n.reason or "No specific reason provided."

            if nominee_id not in grouped_data:
                grouped_data[nominee_id] = {
                    "first_nomination_id": n.id,
                    "name": n.nominee.username,
                    "email": getattr(n.nominee, 'email', 'N/A'),
                    "details": [],
                    "count": 0
                }
            
            detail_str = f"Focus: {cat_str} ({met_str}) -> {reason}"
            grouped_data[nominee_id]["details"].append(detail_str)
            grouped_data[nominee_id]["count"] += 1

        data_for_ai = []
        for nominee_id, data in grouped_data.items():
            combined_text = " | ".join(data["details"])
            full_prompt_text = f"Candidate: {data['name']}. Received {data['count']} nominations. Inputs: {combined_text}"
            data_for_ai.append({
                "id": data["first_nomination_id"], 
                "reason": full_prompt_text
            })

        ai_results = get_nomination_sentiment(data_for_ai)
        ai_lookup = {item['id']: item for item in ai_results}
        final_response = []

        for nominee_id, data in grouped_data.items():
            lookup_id = data["first_nomination_id"]
            ai_data = ai_lookup.get(lookup_id, {})

            final_response.append({
                "id": lookup_id,
                "name": data["name"],
                "email": data["email"],
                "votes": data["count"], 
                "summary": ai_data.get('summary', "Analysis Pending..."),
                "sentiment": ai_data.get('sentiment', "Neutral"),
                "status": "NOMINATION_SUBMITTED" 
            })

        return Response(final_response, status=status.HTTP_200_OK)

class StarAwardExportView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        if request.user.role not in ['ADMIN', 'COORDINATOR']:
            return Response({"error": "Unauthorized"}, status=403)

        nominations = Nomination.objects.filter(
            status__in=[
                'NOMINATION_SUBMITTED',   
                'COORDINATOR_APPROVED', 
                'COORDINATOR_REJECTED',  
                'COMMITTEE_APPROVED', 
                'COMMITTEE_REJECTED',     
                'AWARDED'
            ]
        ).select_related('nominator', 'nominee')

        if not nominations.exists():
            return Response({"error": "No data found to export"}, status=404)

        return generate_star_award_excel(nominations)